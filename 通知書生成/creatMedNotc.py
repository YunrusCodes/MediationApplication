import os
import openpyxl
import shutil
# import PyPDF2
# from docx2pdf import convert # pip install docx2pdf
# from openpyxl import load_workbook
import win32com.client
import pandas as pd
import traceback


def excel_to_pdf(excel_path, pdf_path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.Workbooks.Open(excel_path)
    wb = excel.ActiveWorkbook
    ws = wb.ActiveSheet
    ws.ExportAsFixedFormat(0, pdf_path)
    wb.Close()
    excel.Quit()

def getTwRegion(address):
   if '鄉' in address: 
     return address.split('鄉')[0] + '鄉'
   if '鎮' in address: 
     return address.split('鎮')[0] + '鎮'
   if '區' in address: 
     return address.split('區')[0] + '區'
   if '島' in address: 
     return address.split('島')[0] + '島'
   if '釣魚臺釣魚臺' in address: 
     return '釣魚臺釣魚臺'
   if '釣魚台釣魚台' in address: 
     return '釣魚台釣魚台'
   if '釣魚臺' in address: 
     return address.split('釣魚臺')[0] + '釣魚臺'
   if '釣魚台' in address: 
     return address.split('釣魚台')[0] + '釣魚台'
   if '市' in address: 
     return address.split('市')[0] + '市'

import sys
if getattr(sys, 'frozen', False):
  dir_path = os.path.dirname(sys.executable)
else:
  dir_path = os.path.dirname(os.path.abspath(__file__))

# 讀取檔案
workbook = openpyxl.load_workbook(os.path.join(dir_path, "通知書(空白).xlsx"))
info_book = os.path.join(dir_path, "受通知人資料整理於此.xlsx")

# 新增建立 output 資料夾的功能
output_folder = os.path.join(dir_path, "通知書輸出於此")
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# 創建事件編號表的字典
event_dict = {}
df = pd.read_excel(info_book, sheet_name='通知書資料').fillna('')
event_list = df['事件編號表'].tolist()
for event in event_list:
   if '-' in event:
     event_dict[event.split("-")[0]] = event.split("-")[1]    

# 創建郵遞區號表的字典
postalCode_dict = {}
df = pd.read_excel(info_book, sheet_name='郵遞區號').fillna('')
# 將 DataFrame 轉換為以行政區域為鍵的字典
postalCode_dict = df.set_index('行政區域')['郵遞區號'].to_dict()
addDict = []
for key, value in postalCode_dict.items():
    if '臺' in key: # 台南市、台北市
        region = key.replace('臺','台')
        addDict.append((region, value))
for newDict in addDict:
    postalCode_dict.update([newDict])


    
#生成通知書同時更新sql
import pymysql
from sqlalchemy import create_engine
try:
  # 資料庫連線設定
  host = 'localhost'
  port = 3306
  user = sys.argv[1] if len(sys.argv) >= 2 else input("請輸入使用者名稱：")
  password = sys.argv[2] if len(sys.argv) >= 2 else input("請輸入密碼：")
  database = 'mediation'

  # 連接到 MySQL 資料庫
  conn = pymysql.connect(host=host, port=port, user=user, password=password)
  cursor = conn.cursor()
  cursor.execute(f"CREATE DATABASE IF NOT EXISTS {database}")
  conn = pymysql.connect(host=host, port=port, user=user, password=password,database = database)
  cursor = conn.cursor()
  # 建立 SQLAlchemy 引擎
  engine = create_engine(f'mysql+pymysql://{user}:{password}@{host}:{port}/{database}')

  cursor.execute(f"SHOW TABLES FROM {database} LIKE 'participant'")
  if not cursor.fetchone():
      cursor.execute("use mediation")
      cursor.execute("""
      CREATE TABLE IF NOT EXISTS participant(
          `participant_name` varchar(100),
          `case_num` varchar(100),
          `address` VARCHAR(70));
      """)

  # #調解應攜帶證件(Mediation should bring documents)
  # msbd = os.path.join(dir_path,'額外添加至輸出資料夾中的檔案','08.調解應攜帶證件（欠缺證件者恕無法辦理）.docx')
  # #債權轉讓同意書(Letter of Consent to Assignment of Claims)
  # lc2aoc = os.path.join(dir_path,'額外添加至輸出資料夾中的檔案','07.債權讓與同意書.docx')
  # #委任書(power of attorney)
  # poa = os.path.join(dir_path,'額外添加至輸出資料夾中的檔案','06.委任書.docx')
  # # 轉換 docx 檔案為 PDF 格式
  # pdf_msbd = os.path.join(dir_path, "msbd.pdf")
  # pdf_lc2aoc = os.path.join(dir_path, "lc2aoc.pdf")
  # pdf_poa = os.path.join(dir_path, "poa.pdf")
  # convert(msbd, pdf_msbd)
  # convert(lc2aoc, pdf_lc2aoc)
  # convert(poa, pdf_poa)

# -------------------------------------------------------------------------------------------------------------------------
  # 讀取檔案
  df = pd.read_excel(info_book, sheet_name='業務總覽').fillna('')
  postDate = df['發文日期'].values[0]

  attendanceTime = df['應到時間'].values[0]
  #112年06月06日(星期二)上午09時10分
  med_Date_year = attendanceTime.split('年')[0]
  #attendanceTime.split('年')[1] == 06月06日(星期二)上午09時10分
  med_Date_month = attendanceTime.split('年')[1].split('月')[0]
  #attendanceTime.split('年')[1].attendanceTime.split('月')[1] == 06日(星期二)上午09時10分
  med_Date_day = attendanceTime.split('年')[1].split('月')[1].split('日')[0]

  organiser = df['經辦人'].values[0]
  print(postDate,attendanceTime,organiser)

  # 讀取 Excel 檔案
  df = pd.read_excel(info_book, sheet_name='通知書資料').fillna('')

  # 获取收件編號的所有数字之順序
  receiveNum_order = []
  for i in df['收件編號']:
    if not f"{i:03d}" in receiveNum_order:
      receiveNum_order.append(f"{i:03d}")

  print(receiveNum_order)

  df = pd.read_excel(info_book, sheet_name='通知書資料').fillna('') # 讀取 Excel 檔案
  # 將 DataFrame 轉換為以受通知人為鍵的字典
  notc_dict = df.groupby('受通知人').apply(lambda x: x.drop('受通知人', axis=1).to_dict(orient='records')).to_dict() #同人多案
  df = pd.read_excel(info_book, sheet_name='詳細資料').fillna('')
  # 将DataFrame转换为以受通知人为键的字典
  detail_dict = df.groupby('案號').apply(lambda x: x.drop('案號', axis=1).to_dict(orient='records')).to_dict() #同案多人

  main_dict = {}
  # 重新整合字典，避免重複鍵
  for person, notc_infos in notc_dict.items(): #遍歷所有notc_dict
      for notc_info in notc_infos:          #某個人名涉略哪些案件
          for detail in detail_dict[notc_info['案號']]:
            if detail['受通知人'] == person:
              main_dict[notc_info['案號']+'/'+person] = [ notc_info, detail]  # 添加额外的值到现有键的列表中
              
  #建立一個'調解業務-調解日期'的資料夾 
  collectFolder = os.path.join(output_folder,f'{med_Date_year}年度調解業務-{med_Date_year}.{med_Date_month}.{med_Date_day}-{organiser}')
  # 確保新資料夾不存在，如果存在則刪除
  if os.path.exists(collectFolder):
      shutil.rmtree(collectFolder)
  #再將輸出改變到新增的資料夾 
  os.makedirs(collectFolder)
  output_folder = collectFolder
  shutil.copy2(info_book, output_folder)
  shutil.copy2(os.path.join(dir_path,'mergePDF.exe'), output_folder)


# 由此行開始進行個別資料操作--------------------------------------------------------------------------------------------
  for key, value in main_dict.items():
      recipient = key.split('/')[1]
      if not recipient:
          break
      receiveNum = value[0]['收件編號']  # 收件編號
      address = value[0]['地址']  # 地址
      applicant = value[0]['聲請人'] # 聲請人
      opponents = value[0]['對造人'] # 對造人
      caseNum = key.split('/')[0] # 案件編號
      eventNum = value[0]['事件編號'] # 事件編號
      complete_case_num = caseNum[:3] + '年' + caseNum[3] + '調字第' + caseNum[-3:] + '號'
    
      userQuest = ''
      while not recipient in applicant and not recipient in opponents:
        print(f'收件人-{recipient} 不在此案中 : 聲請人-{applicant} 對造人-{opponents}')

        userQuest = input('是否跳過處裡?[是:輸入"Y" / 否:鍵入ENTER]\n')
        if userQuest == 'y' or userQuest == 'Y':
           break

        userQuest = input('修改收件人?[是:輸入新名稱 / 否:鍵入ENTER]\n')
        if userQuest != '':
           recipient = userQuest

        userQuest = input('修改聲請人?[是:輸入新名稱 / 否:鍵入ENTER]\n')
        if userQuest != '':
           applicant = userQuest
        
        userQuest = input('修改對造人?[是:輸入新名稱 / 否:鍵入ENTER]\n')
        if userQuest != '':
           opponents = userQuest
        userQuest = ''

      if userQuest != '':
         continue

      #更新sql
      cursor.execute(f"SELECT * FROM participant WHERE participant_name = '{recipient}' AND case_num = '{complete_case_num}'")
      if cursor.fetchone():
        cursor.execute(f"UPDATE participant SET address = '{address}' WHERE participant_name = '{recipient}' AND case_num = '{complete_case_num}'")
      else:
        cursor.execute(f"INSERT INTO participant (participant_name, case_num, address) VALUES ( '{recipient}', '{complete_case_num}', '{address}')")


      # 設定目標資料夾路徑
      order = receiveNum_order.index(f"{receiveNum:03d}") + 1
      target_folder = os.path.join(output_folder, f"{order}.調解筆錄{receiveNum:03d}(聲請人{applicant})(對造人{opponents})(收：{caseNum})開調解時間{attendanceTime})-({event_dict[eventNum]})")
      if not os.path.exists(target_folder) :
        os.makedirs(target_folder)
      
      
      # 取得工作表
      sheet = workbook[workbook.sheetnames[0]]
      sheet['B3'].value = postDate 
      sheet['B9'].value = attendanceTime
      sheet['B6'].value = recipient
      try:
        sheet['B4'].value = str(postalCode_dict[getTwRegion(address)]) + address
      except KeyError as e:
        address = input(f'收件人 {recipient} 地址錯誤:\n{address}\n請更正:\n')
        try:
          sheet['B4'].value = str(postalCode_dict[getTwRegion(address)]) + address
        except:
          print(f'收件人 {recipient} 修正地址錯誤:{address} 無法添加郵遞區號')
          sheet['B4'].value = address
      sheet['B7'].value = complete_case_num
      sheet['B8'].value = applicant + "與" + opponents + "間「" + event_dict[eventNum] + "」"

      filename = ""
      # 受通知人是聲請人
      if recipient == applicant:
        filename = "01-1" + '.' + "調解通知書" + "-" + recipient + ".xlsx"
      else:   #受通知人是對造人之一
        for i in range(len(opponents.split("、"))):
          if recipient == opponents.split("、")[i]:
            filename = "01-" + str(i+2) + '.' + "調解通知書" + "-" + recipient + ".xlsx"

      print(filename)
      workbook.save(os.path.join(output_folder,target_folder,filename))

      filesToAppend = os.path.join(dir_path,'額外添加至輸出資料夾中的檔案')
      # 將目錄中的檔案複製到目標資料夾中
      for file in os.listdir(filesToAppend):
          file_path = os.path.join(filesToAppend, file)
          if os.path.isfile(file_path):
            try:
              shutil.copy(file_path, target_folder)
            except FileExistsError:
              print(f"檔案已存在{file_path}")
            except Exception as e:
              print(f"執行失敗{e}")
      




  #     #調解通知書
  #     medNotice = output_path
  #     pdf_medNotc = os.path.join(target_folder, "medNotc.pdf")
  #     excel_to_pdf(output_path,pdf_medNotc)

  #     # 將 xlsx 和兩個 PDF 合併成一個 PDF
  #     pdf_writer = PyPDF2.PdfWriter()

  #     with open(pdf_msbd, 'rb') as pdf1, open(pdf_lc2aoc, 'rb') as pdf2, open(pdf_poa, 'rb') as pdf3, open(pdf_medNotc, 'rb') as pdf4:
  #         pdf_reader1 = PyPDF2.PdfReader(pdf1)
  #         pdf_reader2 = PyPDF2.PdfReader(pdf2)
  #         pdf_reader3 = PyPDF2.PdfReader(pdf3)
  #         excel_reader = PyPDF2.PdfReader(pdf4)
          
  #         pdf_writer.add_page(excel_reader.pages[0])
  #         pdf_writer.add_page(pdf_reader1.pages[0])
  #         pdf_writer.add_page(pdf_reader2.pages[0])
  #         pdf_writer.add_page(pdf_reader3.pages[0])

  #     fileToPrint = os.path.join(target_folder, f"待印通知書({recipient}).pdf")
  #     os.remove(pdf_medNotc)
  #     # 儲存 PDF 檔案
  #     with open(fileToPrint, 'wb') as out:
  #         pdf_writer.write(out)


  # os.remove(pdf_msbd)
  # os.remove(pdf_lc2aoc)
  # os.remove(pdf_poa)
  # 關閉檔案
  workbook.close()

  #提交並關閉sql
  conn.commit()
  conn.close()  


except Exception as e:
   print('發生錯誤:')
   traceback.print_exc()
   input(f"請按任意鍵離開")
   sys.exit()
