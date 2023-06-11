import os
import unicodedata

def to_halfwidth(s):
    # 把全形字轉成半形字
    return unicodedata.normalize('NFKC', s)

def count_lines(s, w):
    line_count = 0
    word_count = 0
    for char in s:
        # 判斷是否為中文字或中文標點符號
        if '\u4e00' <= char <= '\u9fff' or char in '，。！？；：、（）『』「」《》【】':
            word_count += 1
        # 判斷是否為數字或其他非中文符號
        elif char.isdigit() or char in '.,;?!':
            word_count += 0.5
        # 判斷是否為換行符號
        elif char == '\n':
            line_count += 1
            word_count = 0
        # 忽略其他非中文字符
        else:
            pass
        
        # 如果超過指定的字數，換行計數
        if word_count > w:
            line_count += 1
            word_count = 0
    
    # 如果還有未滿一行的字，計為一行
    if word_count > 0:
        line_count += 1
    
    return line_count

class Case:
    def __init__(self):
        self.number = None             # 收件編號
        self.time = None               # 收件日期
        self.case_number = None        # 案號
        self.applicants = ""           # 聲請人
        self.opponents = ""            # 對造人
        self.case_reason = None        # 事由
        self.summary = ""              # 過程摘要
        self.mediation_result = None   # 調解結果
        self.committee = None          # 委員
        self.review_date = None        # 報院審查日期
        self.review_number = None      # 報院審查文號
        self.return_date = None        # 法院發還日期
        self.return_number = None      # 法院發還文號
        self.review_result = None      # 法院審查結果
        self.transfer_agency = None    # 轉借機關
        self.referrer = None           # 轉介人
        self.filing_date = None        # 歸檔日期

import openpyxl
from openpyxl.styles import Font
# 輸入檔案路徑
import sys
if getattr(sys, 'frozen', False):
    absPath = os.path.dirname(sys.executable)
else:
    absPath = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(absPath, '空白調解進行簿.xlsx')



import pymysql
from sqlalchemy import create_engine
import re

# 資料庫連線設定
host = 'localhost'
port = 3306
user = sys.argv[1] if len(sys.argv) >= 2 else input("請輸入使用者名稱：")
password = sys.argv[2] if len(sys.argv) >= 2 else input("請輸入密碼：")
database = 'mediation'
try:
  # 連接到 MySQL 資料庫
  conn = pymysql.connect(host=host, port=port, user=user, password=password,database=database)
  cursor = conn.cursor()
except Exception as e:
   input(f"資料庫連接失敗:{e}\n請按任一鍵離開")
   sys.exit()

try:
  cursor.execute(f"SELECT * FROM {database}.med_case")
  column_names = [column[0] for column in cursor.description]
  datas = cursor.fetchall()
  results = []
  for data in datas:
      result = {}
      for i in range(len(column_names)):
          result[column_names[i]] = data[i]
      results.append(result)

  data_list = []
  for result in results:
      
      cursor.execute(f"SELECT * FROM {database}.contents")
      column_names = [column[0] for column in cursor.description]
      
      contents_data = {}
      cursor.execute(f"SELECT * FROM {database}.contents WHERE `案號`='{result['案號']}'")
      data = cursor.fetchone()
      if data:
        for i in range(len(column_names)):
            contents_data[column_names[i]] = data[i]
      else:
        for i in range(len(column_names)):
            contents_data[column_names[i]] = None   
      
      data_dict = {
          '收件編號': result['收件編號'],
          '收件日期': result['收件日期'],
          '案號': result['案號'],
          '聲請人': result['聲請人'],
          '對造人': result['對造人'],
          '事由': result['事由'],
          '過程摘要': result['過程摘要'],

          '轉介單位': contents_data['轉介單位'] if contents_data else None,
          '轉介人': contents_data['轉介人'] if contents_data else None,
          '委任人': contents_data['委任人'] if contents_data else None,
          '受任人': contents_data['受任人'] if contents_data else None,
          '報院審查日期': contents_data['報院審查日期'] if contents_data else None,
          '法院發還日期': contents_data['法院發還日期'] if contents_data else None,
          '法院審查結果': contents_data['法院審查結果'] if contents_data else None,
          '調解委員': contents_data['調解委員'] if contents_data else None,
          '主席': contents_data['主席'] if contents_data else None,
          '法院審查結果': contents_data['法院審查結果'] if contents_data else None,
          '調解結果' :  '私下和解' if contents_data['私下和解註記'] != None and contents_data['私下和解註記'].strip() != None 
                                  else '調解不成立' if contents_data['調解書'] == 0 or contents_data['調解書'] == None
                                  else '調解成立',
          '最後送達日期': contents_data['最後送達日期'] if contents_data else None,
          '私下和解註記': contents_data['私下和解註記'] if contents_data else None,
          '報院審查文號': contents_data['報院審查文號'] if contents_data else None,
          '法院發還文號': contents_data['法院發還文號'] if contents_data else None,
          '最後送達日期': contents_data['最後送達日期'] if contents_data else None,
      }
      print('--------------',contents_data['收件編號'],contents_data['調解書'],data_dict['調解結果'])
      data_list.append(data_dict)

  # for data in data_list:
  #     for key, value in data.items():
  #         print(key,"->",value)

  # 讀取空白調解進行簿.xlsx的workbook
  workbook = openpyxl.load_workbook(file_path)
  # 輸出檔案的路徑
  output_path = os.path.join(absPath, '解析結果輸出於此', '輸出調解進行簿.xlsx')

  # workbook.save(output_path)
  source_worksheet = workbook.worksheets[0]

  max_number = 0
  # 找出最大的 value
  for data in data_list:
    if max_number < int(data['收件編號']) :
        max_number = int(data['收件編號'])

  print('最大數' + str(max_number))
  # 計算每個 workbook 的數量
  num_workbooks = int(max_number / 3)

  # 複製第一個 workbook num_workbooks 次
  for i in range(2,num_workbooks+2):
      new_worksheet = workbook.copy_worksheet(source_worksheet)
      new_worksheet.title = str(i*3-2) +'-'+ str(i*3)
  # 儲存 workbook
  workbook.save(output_path)
  print('最大'+str(max_number))
  print('有'+str(num_workbooks))
  # 选择第十个 worksheet，并将 "HELLO" 写入 C1 单元格
  # target_worksheet = workbook.worksheets[9]
  # target_worksheet['C1'] = "HELLO"

  for data_dict in data_list:

    #指定workbook
    target_worksheet = workbook.worksheets[(int(data_dict['收件編號'])-1)//3] #1-0 2-0 3-0 4-1 5-1 6-1
    #依收件編號指定column
    col_num = (int(data_dict['收件編號'])-1) % 3  #1-0 2-1 3-2 4-1 5-1 6-1

    #寫入收件編號
    target_cell = chr(ord('C') + col_num ) + '1'
    target_worksheet[target_cell] = data_dict['收件編號']

    #寫入收件日期
    if data_dict['收件日期'] != None:
      target_cell = chr(ord('C') + col_num ) + '2'
      target_worksheet[target_cell] = data_dict['收件日期'].split('時')[0]+'時' if '時' in data_dict['收件日期'] else data_dict['收件日期']

    #寫入案號
    if data_dict['案號'] != None:
      target_cell = chr(ord('C') + col_num ) + '3'
      target_worksheet[target_cell] = data_dict['案號'].split('字')[0]+'字\n' + data_dict['案號'].split('字')[1]

    #寫入聲請人姓名
    if data_dict['聲請人'] != None:
      target_cell = chr(ord('C') + col_num ) + '4'
      target_worksheet[target_cell] = data_dict['聲請人'].replace(",","、")

    #寫入對造人姓名
    if data_dict['對造人'] != None:
      target_cell = chr(ord('C') + col_num ) + '5'
      target_worksheet[target_cell] = data_dict['對造人'].replace(",","、")

    #寫入事由
    if data_dict['事由'] != None:
      target_cell = chr(ord('C') + col_num ) + '6'
      target_worksheet[target_cell] = data_dict['事由']
      
    #寫入過程摘要
    if data_dict['過程摘要']:
      sum_lines = count_lines(data_dict['過程摘要'].split('\n')[1:],13)
      target_cell = chr(ord('C') + col_num ) + '7'
      target_worksheet[target_cell] =  data_dict['過程摘要']
      if sum_lines < 16:
        target_worksheet[target_cell].font = Font(name='標楷體',size= 9 )
      else:
        target_worksheet[target_cell].font = Font(name='標楷體',size= 9 - ((sum_lines)-16)/2 )

    #寫入調解結果
    if data_dict['調解結果'] != None:
      target_cell = chr(ord('C') + col_num ) + '10'
      target_worksheet[target_cell] = data_dict['調解結果']
    
    #寫入調解委員
    if data_dict['調解委員'] != None:
      target_cell = chr(ord('C') + col_num ) + '16'
      target_worksheet[target_cell] = '獨任調解/' + data_dict['調解委員']

    #寫入報院審查日期
    if data_dict['報院審查日期'] != None:
      target_cell = chr(ord('C') + col_num ) + '11'
      target_worksheet[target_cell] = data_dict['報院審查日期']

    #寫入法院發還日期
    if data_dict['法院發還日期'] != None:
      target_cell = chr(ord('C') + col_num ) + '12'
      target_worksheet[target_cell] = data_dict['法院發還日期']

    #寫入法院發還日期
    if data_dict['法院審查結果'] != None:
      target_cell = chr(ord('C') + col_num ) + '13'
      target_worksheet[target_cell] = data_dict['法院審查結果']

    #寫入最後送達日期
    if data_dict['最後送達日期'] != None:
      target_cell = chr(ord('C') + col_num ) + '14'
      target_worksheet[target_cell] = data_dict['最後送達日期']
    
    #寫入轉介單位與轉介人
    writeRefer = ''
    if data_dict['轉介單位'] != None:
      if '警察局' in data_dict['轉介單位'] and '分局' in data_dict['轉介單位']:
        if data_dict['轉介單位'].split('分局')[1]:
          writeRefer = data_dict['轉介單位'].split('分局')[1] 
        else:
          writeRefer = data_dict['轉介單位'].split('警察局')[1].split('分局')[0] + '交通分隊'
      else:
        writeRefer = data_dict['轉介單位']
      if data_dict['轉介人'] != None:
        writeRefer += '/' + data_dict['轉介人']
      target_cell = chr(ord('C') + col_num ) + '18'
      target_worksheet[target_cell] = writeRefer

      # case.review_number = worksheet.cell(row=row_num, column=11).value
      # case.return_date = worksheet.cell(row=row_num, column=12).value
      # case.return_number = worksheet.cell(row=row_num, column=13).value
      # case.review_result = worksheet.cell(row=row_num, column=14).value
      # case.transfer_agency = worksheet.cell(row=row_num, column=15).value
      # case.referrer = worksheet.cell(row=row_num, column=16).value
      # case.filing_date = worksheet.cell(row=row_num, column=17).value

  # 儲存 workbook
  workbook.save(output_path)

except Exception as e:
  input(f"發生錯誤:{e}\n請按任一鍵離開")
  sys.exit()   